from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, Count
from django.utils import timezone
from django.http import HttpResponse

from .utils import montant_en_lettres_fr

from .forms import LoginForm, PatientForm, AppointmentForm, PaymentForm, PaymentItemForm, ExpenseForm, LabPackApplyForm, SupplierForm, PurchaseForm, PurchaseItemForm, EmployeeForm, SalaryForm, MedecinAmbulantForm, HonorairesPayerForm
from .models import Patient, Appointment, Payment, PaymentItem, Expense, LabTest, LabPack, ServiceType, AuditLog, Supplier, Purchase, PurchaseItem, Employee, Salary, PurchaseStatus, TarifAccouchement, AccouchementDetail, Doctor, MedecinAmbulant
from .permissions import require_groups, GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN, in_groups


def _next_dossier_numero():
    """Génère un numéro automatique : DOS-YYYY-NNNN"""
    from django.utils import timezone
    year = timezone.now().year
    prefix = f"DOS-{year}-"
    last = (DossierMaternite.objects
            .filter(numero__startswith=prefix)
            .order_by("-numero")
            .values_list("numero", flat=True)
            .first())
    if last:
        try:
            n = int(last.split("-")[-1]) + 1
        except Exception:
            n = 1
    else:
        n = 1
    return f"{prefix}{n:04d}"


def _next_examen_numero(type_examen):
    """Génère un numéro automatique : EX-TYPE-YYYY-NNNN"""
    from django.utils import timezone
    year = timezone.now().year
    prefix = f"EX-{type_examen}-{year}-"
    last = (ExamenMaternite.objects
            .filter(numero_manuel__startswith=prefix)
            .order_by("-numero_manuel")
            .values_list("numero_manuel", flat=True)
            .first())
    if last:
        try:
            n = int(last.split("-")[-1]) + 1
        except Exception:
            n = 1
    else:
        n = 1
    return f"{prefix}{n:04d}"

def login_view(request):
    if request.user.is_authenticated:
        return redirect("home")
    form = LoginForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        login(request, form.get_user())
        return redirect("home")
    return render(request, "auth/login.html", {"form": form})

def logout_view(request):
    logout(request)
    return redirect("login")

@login_required
def home(request):
    groups = set(request.user.groups.values_list("name", flat=True))
    today = timezone.localdate()
    kpi = {
        "rdv_today": Appointment.objects.filter(start_at__date=today).count(),
        "payments_today": Payment.objects.filter(paid_at__date=today).aggregate(s=Sum("amount_total"))["s"] or 0,
        "patients_total": Patient.objects.count(),
    }
    next_rdv = Appointment.objects.filter(start_at__date=today).select_related("patient")[:12]
    return render(request, "home.html", {"kpi": kpi, "next_rdv": next_rdv})

# Patients
@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def patient_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = Patient.objects.all()
    if q:
        qs = qs.filter(phone__icontains=q) | qs.filter(first_name__icontains=q) | qs.filter(last_name__icontains=q)
    today = timezone.localdate()
    month_start = today.replace(day=1)
    stats = {
        "total": Patient.objects.count(),
        "new_month": Patient.objects.filter(created_at__date__gte=month_start).count(),
        "new_today": Patient.objects.filter(created_at__date=today).count(),
    }
    return render(request, "patients/list.html", {"patients": qs[:300], "q": q, "stats": stats})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def patient_new(request):
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        last_name = request.POST.get("last_name","").strip()
        first_name = request.POST.get("first_name","").strip()
        phone = request.POST.get("phone","").strip()
        birth_date = request.POST.get("birth_date","").strip() or None
        notes = request.POST.get("notes","").strip()
        if not last_name: errors["last_name"] = "Champ obligatoire"
        if not first_name: errors["first_name"] = "Champ obligatoire"
        if not phone: errors["phone"] = "Champ obligatoire"
        if not errors:
            try:
                p = Patient.objects.create(last_name=last_name, first_name=first_name,
                    phone=phone, birth_date=birth_date or None, notes=notes)
                AuditLog.objects.create(actor=request.user, action="CREATE", entity="Patient", entity_id=str(p.id), message=str(p))
                messages.success(request, "Patient ajouté.")
                return redirect("patient_detail", pk=p.id)
            except Exception as e:
                errors["phone"] = "Ce numéro existe déjà."
    return render(request, "patients/form.html", {"errors": errors, "vals": vals, "title": "Nouveau patient"})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def patient_detail(request, pk:int):
    p = get_object_or_404(Patient, pk=pk)
    return render(request, "patients/detail.html", {"p": p})

# RDV
@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def appointment_list(request):
    import datetime
    today = timezone.localdate()
    day = request.GET.get("day")
    if day:
        try:
            day = timezone.datetime.fromisoformat(day).date()
        except Exception:
            day = today
    else:
        day = today
    prev_day = day - datetime.timedelta(days=1)
    next_day = day + datetime.timedelta(days=1)
    qs = Appointment.objects.filter(start_at__date=day).select_related("patient", "doctor").order_by("start_at")
    stats = {
        "confirmed": qs.filter(status="CONFIRMED").count(),
        "done": qs.filter(status="DONE").count(),
        "cancelled": qs.filter(status="CANCELLED").count(),
        "pending": qs.filter(status="PENDING").count(),
    }
    return render(request, "rdv/list.html", {
        "appointments": qs, "day": day,
        "prev_day": prev_day, "next_day": next_day,
        "today": today, "stats": stats,
    })

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def appointment_new(request):
    from .models import Doctor, ServiceType, AppointmentStatus
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        patient_id = request.POST.get("patient_id","").strip()
        start_at = request.POST.get("start_at","").strip()
        service_type = request.POST.get("service_type","").strip()
        status = request.POST.get("status","CONFIRMED")
        doctor_id = request.POST.get("doctor_id","").strip()
        note = request.POST.get("note","").strip()
        if not patient_id: errors["patient_id"] = "Choisir un patient"
        if not start_at: errors["start_at"] = "Date obligatoire"
        if not service_type: errors["service_type"] = "Type obligatoire"
        if not errors:
            try:
                rdv = Appointment(service_type=service_type, status=status, note=note, created_by=request.user)
                rdv.patient_id = int(patient_id)
                if doctor_id: rdv.doctor_id = int(doctor_id)
                from django.utils.dateparse import parse_datetime
                rdv.start_at = parse_datetime(start_at.replace("T"," ") + ":00") or parse_datetime(start_at)
                rdv.save()
                AuditLog.objects.create(actor=request.user, action="CREATE", entity="Appointment", entity_id=str(rdv.id), message=str(rdv))
                messages.success(request, "RDV créé.")
                return redirect("appointment_list")
            except Exception as e:
                errors["general"] = str(e)
    patients = Patient.objects.order_by("last_name","first_name")
    doctors = Doctor.objects.all()
    return render(request, "rdv/form.html", {
        "errors": errors, "vals": vals, "title": "Nouveau RDV",
        "patients": patients, "doctors": doctors,
        "service_types": ServiceType.choices,
        "statuses": AppointmentStatus.choices,
    })

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def appointment_edit(request, pk:int):
    from .models import Doctor, ServiceType, AppointmentStatus
    rdv = get_object_or_404(Appointment, pk=pk)
    errors = {}
    if request.method == "POST":
        patient_id = request.POST.get("patient_id","").strip()
        start_at = request.POST.get("start_at","").strip()
        service_type = request.POST.get("service_type","").strip()
        status = request.POST.get("status","CONFIRMED")
        doctor_id = request.POST.get("doctor_id","").strip()
        note = request.POST.get("note","").strip()
        if not patient_id: errors["patient_id"] = "Choisir un patient"
        if not start_at: errors["start_at"] = "Date obligatoire"
        if not errors:
            try:
                rdv.patient_id = int(patient_id)
                rdv.service_type = service_type
                rdv.status = status
                rdv.note = note
                rdv.doctor_id = int(doctor_id) if doctor_id else None
                from django.utils.dateparse import parse_datetime
                rdv.start_at = parse_datetime(start_at.replace("T"," ") + ":00") or parse_datetime(start_at)
                rdv.save()
                AuditLog.objects.create(actor=request.user, action="UPDATE", entity="Appointment", entity_id=str(rdv.id), message="Modification RDV")
                messages.success(request, "RDV modifié.")
                return redirect("appointment_list")
            except Exception as e:
                errors["general"] = str(e)
    patients = Patient.objects.order_by("last_name","first_name")
    doctors = Doctor.objects.all()
    return render(request, "rdv/form.html", {
        "errors": errors, "rdv": rdv, "title": "Modifier RDV",
        "patients": patients, "doctors": doctors,
        "service_types": ServiceType.choices,
        "statuses": AppointmentStatus.choices,
    })

# Payments

def _barcode_url(receipt_no: str) -> str:
    """Génère un code-barres CODE128 dans static/barcodes et retourne l'URL /static/..."""
    try:
        import os
        from django.conf import settings
        import barcode
        from barcode.writer import ImageWriter
        folder = os.path.join(settings.BASE_DIR, "static", "barcodes")
        os.makedirs(folder, exist_ok=True)
        base_path = os.path.join(folder, receipt_no)
        png_path = base_path + ".png"
        if not os.path.exists(png_path):
            ean = barcode.get("code128", receipt_no, writer=ImageWriter())
            ean.save(base_path)  # creates .png
        return f"/static/barcodes/{receipt_no}.png"
    except Exception:
        return ""


def _receipt_no():
    # Simple receipt number: YYYYMMDD-XXXXX
    now = timezone.now()
    prefix = now.strftime("%Y%m%d")
    import random
    import string
    return f"{prefix}-" + "".join(random.choices(string.digits, k=5))

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def payment_list(request):
    q = (request.GET.get("q") or "").strip()
    date_from = request.GET.get("date_from") or ""
    date_to   = request.GET.get("date_to") or ""
    payer     = request.GET.get("payer") or ""
    qs = Payment.objects.select_related("patient", "doctor").all()
    if q:
        qs = qs.filter(receipt_no__icontains=q) | qs.filter(patient__phone__icontains=q) | qs.filter(patient__last_name__icontains=q)
    if date_from:
        try: qs = qs.filter(paid_at__date__gte=date_from)
        except: pass
    if date_to:
        try: qs = qs.filter(paid_at__date__lte=date_to)
        except: pass
    if payer:
        qs = qs.filter(payer_type=payer)
    total_filtre = qs.aggregate(s=Sum("amount_total"))["s"] or 0
    today = timezone.localdate()
    total_today = Payment.objects.filter(paid_at__date=today).aggregate(s=Sum("amount_total"))["s"] or 0
    month_start = today.replace(day=1)
    total_month = Payment.objects.filter(paid_at__date__gte=month_start).aggregate(s=Sum("amount_total"))["s"] or 0
    from .models import PayerType
    return render(request, "payments/list.html", {
        "payments": qs[:300], "q": q,
        "date_from": date_from, "date_to": date_to, "payer": payer,
        "total_filtre": total_filtre, "total_today": total_today, "total_month": total_month,
        "PayerType": PayerType,
    })

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def payment_new(request):
    form = PaymentForm(request.POST or None)
    item_form = PaymentItemForm()
    pack_form = LabPackApplyForm()
    lab_tests = LabTest.objects.filter(is_active=True)

    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            pay = form.save(commit=False)
            pay.created_by = request.user
            pay.receipt_no = _receipt_no()
            pay.amount_total = (pay.amount_patient or 0) + (pay.amount_third_party or 0)
            pay.save()

            labels = request.POST.getlist("item_label[]")
            lab_ids = request.POST.getlist("item_lab_test[]")
            amounts = request.POST.getlist("item_amount[]")
            total = pay.amount_total

            if pay.service_type == ServiceType.ANALYSES:
                total = 0
                for i in range(len(amounts)):
                    amt_str = (amounts[i] or "0").replace(",", ".")
                    try:
                        amt = float(amt_str)
                    except Exception:
                        amt = 0
                    lab_id = lab_ids[i] if i < len(lab_ids) else ""
                    label = labels[i] if i < len(labels) else ""
                    lab_obj = None
                    if lab_id:
                        try:
                            lab_obj = LabTest.objects.get(id=int(lab_id))
                        except Exception:
                            lab_obj = None
                    PaymentItem.objects.create(payment=pay, lab_test=lab_obj, label=label, amount=amt)
                    total += amt

            # Accouchement detail
            if pay.service_type in (ServiceType.ACC_NATUREL, ServiceType.CESARIENNE):
                def _d(key, default=0):
                    v = request.POST.get(key, "").strip()
                    try: return float(v) if v else float(default)
                    except: return float(default)
                nb_nuits = int(request.POST.get("acc_nb_nuits","1") or 1)
                tarif_nuit = _d("acc_tarif_nuit")
                med_amb_raw = request.POST.get("acc_medecin_ambulant", "").strip()
                acc = AccouchementDetail(
                    payment=pay,
                    type_acte=pay.service_type,
                    salle=_d("acc_salle"),
                    honoraires_medecin=_d("acc_honoraires_medecin"),
                    nb_nuits=nb_nuits,
                    tarif_nuit=tarif_nuit,
                    anesthesie=_d("acc_anesthesie"),
                    sage_femme=_d("acc_sage_femme"),
                    medicaments=_d("acc_medicaments"),
                    frais_admin=_d("acc_frais_admin"),
                    certificat_naissance=_d("acc_certificat"),
                    frais_chifa=_d("acc_frais_chifa"),
                    notes=request.POST.get("acc_notes","").strip(),
                )
                if med_amb_raw:
                    try:
                        acc.medecin_ambulant_id = int(med_amb_raw)
                    except (ValueError, TypeError):
                        pass
                acc.save()
                total = acc.total_frais()

            pay.amount_total = total
            if pay.payer_type == 'NORMAL':
                pay.amount_patient = total
                pay.amount_third_party = 0
            pay.save()

            AuditLog.objects.create(actor=request.user, action="CREATE", entity="Payment", entity_id=str(pay.id), message=f"Reçu {pay.receipt_no} {pay.amount_total}")
            messages.success(request, f"Encaissement enregistré. Reçu: {pay.receipt_no}")
            return redirect("payment_receipt", pk=pay.id)

    tarifs = TarifAccouchement.get()
    from .models import MedicamentConsommable
    medicaments = MedicamentConsommable.objects.filter(is_active=True).order_by("nom")
    medecins_ambulants = MedecinAmbulant.objects.filter(is_active=True)
    return render(request, "payments/form.html", {
        "form": form,
        "item_form": item_form,
        "pack_form": pack_form,
        "lab_tests": lab_tests,
        "service_type_analyses": ServiceType.ANALYSES,
        "title": "Nouvel encaissement",
        "tarifs": tarifs,
        "medicaments": medicaments,
        "medecins_ambulants": medecins_ambulants,
    })

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def payment_receipt(request, pk:int):
    pay = get_object_or_404(Payment.objects.select_related("patient", "created_by", "doctor"), pk=pk)
    items = pay.items.select_related("lab_test").all()

    montant_lettres = montant_en_lettres_fr(pay.amount_total)
    barcode_url = _barcode_url(pay.receipt_no)

    paid = (pay.amount_patient or 0) + (pay.amount_third_party or 0)
    unpaid = (pay.amount_total or 0) - paid

    try:
        acc_detail = pay.acc_detail
    except Exception:
        acc_detail = None

    return render(request, "payments/receipt_a5.html", {
        "pay": pay,
        "items": items,
        "montant_lettres": montant_lettres,
        "barcode_url": barcode_url,
        "paid": paid,
        "unpaid": unpaid,
        "acc_detail": acc_detail,
    })
# Expenses
@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def expense_list(request):
    cat = request.GET.get("cat") or ""
    date_from = request.GET.get("date_from") or ""
    date_to   = request.GET.get("date_to") or ""
    qs = Expense.objects.select_related("created_by").all()
    if cat:
        qs = qs.filter(category=cat)
    if date_from:
        try: qs = qs.filter(spent_at__gte=date_from)
        except: pass
    if date_to:
        try: qs = qs.filter(spent_at__lte=date_to)
        except: pass
    total_filtre = qs.aggregate(s=Sum("amount"))["s"] or 0
    today = timezone.localdate()
    month_start = today.replace(day=1)
    total_month = Expense.objects.filter(spent_at__gte=month_start).aggregate(s=Sum("amount"))["s"] or 0
    from .models import ExpenseCategory
    return render(request, "expenses/list.html", {
        "expenses": qs[:300], "cat": cat,
        "date_from": date_from, "date_to": date_to,
        "total_filtre": total_filtre, "total_month": total_month,
        "ExpenseCategory": ExpenseCategory,
    })

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def expense_export_excel(request):
    # Export XLSX (Excel)
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Dépenses"

    ws.append(["Date", "Catégorie", "Montant", "Note", "Créé par"])
    for e in Expense.objects.select_related("created_by").order_by("-spent_at")[:5000]:
        ws.append([
            e.spent_at.strftime("%Y-%m-%d"),
            e.get_category_display(),
            float(e.amount),
            e.note or "",
            (e.created_by.username if e.created_by else ""),
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="depenses.xlsx"'
    return resp


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def expense_export_pdf(request):
    # Export PDF
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    y = h - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Gestion des Dépenses")
    y -= 24

    c.setFont("Helvetica", 10)
    total = Expense.objects.aggregate(s=Sum("amount"))["s"] or 0
    c.drawString(40, y, f"Total: {total} DA")
    y -= 18

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Date")
    c.drawString(120, y, "Catégorie")
    c.drawRightString(w - 120, y, "Montant")
    c.drawString(w - 110, y, "Note")
    y -= 14
    c.setFont("Helvetica", 9)

    for e in Expense.objects.order_by("-spent_at")[:200]:
        if y < 60:
            c.showPage()
            y = h - 40
            c.setFont("Helvetica", 9)
        c.drawString(40, y, e.spent_at.strftime("%Y-%m-%d"))
        c.drawString(120, y, e.get_category_display()[:18])
        c.drawRightString(w - 120, y, f"{e.amount}")
        c.drawString(w - 110, y, (e.note or "")[:30])
        y -= 12

    c.showPage()
    c.save()
    bio.seek(0)

    resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="depenses.pdf"'
    return resp

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def expense_new(request):
    from .models import ExpenseCategory
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        category = request.POST.get("category","").strip()
        amount = request.POST.get("amount","").strip()
        spent_at = request.POST.get("spent_at","").strip()
        note = request.POST.get("note","").strip()
        if not category: errors["category"] = "Catégorie obligatoire"
        if not amount: errors["amount"] = "Montant obligatoire"
        if not spent_at: errors["spent_at"] = "Date obligatoire"
        if not errors:
            try:
                e = Expense.objects.create(category=category, amount=float(amount),
                    spent_at=spent_at, note=note, created_by=request.user)
                AuditLog.objects.create(actor=request.user, action="CREATE", entity="Expense", entity_id=str(e.id), message=f"{e.category} {e.amount}")
                messages.success(request, "Dépense enregistrée.")
                return redirect("expense_list")
            except Exception as ex:
                errors["general"] = str(ex)
    return render(request, "expenses/form.html", {
        "errors": errors, "vals": vals,
        "categories": ExpenseCategory.choices,
    })

# Reports

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def expense_edit(request, pk):
    e = get_object_or_404(Expense, pk=pk)
    from .models import ExpenseCategory
    errors = {}
    if request.method == "POST":
        if "delete" in request.POST:
            e.delete()
            messages.success(request, "Dépense supprimée.")
            return redirect("expense_list")
        category  = request.POST.get("category","").strip()
        amount    = request.POST.get("amount","").strip()
        spent_at  = request.POST.get("spent_at","").strip()
        note      = request.POST.get("note","").strip()
        if not amount: errors["amount"] = "Montant obligatoire"
        if not errors:
            try:
                e.category = category; e.amount = float(amount)
                e.spent_at = spent_at; e.note = note; e.save()
                messages.success(request, "Dépense modifiée.")
                return redirect("expense_list")
            except Exception as ex:
                errors["general"] = str(ex)
    return render(request, "expenses/form.html", {
        "errors": errors, "expense": e, "edit": True,
        "categories": ExpenseCategory.choices,
    })


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def caisse_solde(request):
    """Vue tableau de bord caisse — solde en temps réel."""
    from django.utils import timezone
    today = timezone.localdate()

    # Entrées : tous les encaissements
    entrees = Payment.objects.aggregate(s=Sum("amount_total"))["s"] or 0

    # Sorties
    depenses = Expense.objects.aggregate(s=Sum("amount"))["s"] or 0
    salaires = Salary.objects.aggregate(s=Sum("amount"))["s"] or 0
    achats   = Purchase.objects.aggregate(s=Sum("total_amount"))["s"] or 0

    sorties  = depenses + salaires + achats
    solde    = entrees - sorties

    # Dernières opérations (journal)
    from itertools import chain
    import operator

    ops = []
    for p in Payment.objects.select_related("patient").order_by("-paid_at")[:20]:
        ops.append({
            "date":    p.paid_at,
            "libelle": f"Encaissement — {p.patient.last_name} {p.patient.first_name}",
            "type":    "entree",
            "montant": float(p.amount_total),
            "ref":     p.receipt_no,
        })
    for e in Expense.objects.select_related("created_by").order_by("-spent_at")[:20]:
        ops.append({
            "date":    e.spent_at,
            "libelle": f"Dépense — {e.get_category_display()} {e.note or ''}".strip(),
            "type":    "sortie",
            "montant": float(e.amount),
            "ref":     f"DEP-{e.id}",
        })
    for s in Salary.objects.select_related("employee").order_by("-month")[:20]:
        ops.append({
            "date":    s.month,
            "libelle": f"Salaire — {s.employee.full_name}",
            "type":    "sortie",
            "montant": float(s.amount),
            "ref":     f"SAL-{s.id}",
        })
    for a in Purchase.objects.select_related("supplier").order_by("-purchased_at")[:10]:
        ops.append({
            "date":    a.purchased_at,
            "libelle": f"Achat — {a.supplier.name}",
            "type":    "sortie",
            "montant": float(a.total_amount),
            "ref":     a.invoice_no or f"ACH-{a.id}",
        })

    ops.sort(key=lambda x: str(x["date"]), reverse=True)

    return render(request, "finance/caisse_solde.html", {
        "solde": solde, "entrees": entrees, "sorties": sorties,
        "depenses": depenses, "salaires": salaires, "achats": achats,
        "ops": ops[:40],
        "today": today,
    })

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def reports(request):
    import datetime
    today = timezone.localdate()
    # Période filtrable
    date_from_str = request.GET.get("date_from","")
    date_to_str   = request.GET.get("date_to","")
    try:
        date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else today.replace(day=1)
    except Exception:
        date_from = today.replace(day=1)
    try:
        date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else today
    except Exception:
        date_to = today

    # Encaissements
    pays_qs = Payment.objects.filter(paid_at__date__gte=date_from, paid_at__date__lte=date_to).select_related("patient","doctor")
    pay_total   = pays_qs.aggregate(s=Sum("amount_total"))["s"] or 0
    pay_chifa   = pays_qs.filter(payer_type="CHIFFA").aggregate(s=Sum("amount_third_party"))["s"] or 0
    pay_mil     = pays_qs.filter(payer_type="MIL").aggregate(s=Sum("amount_third_party"))["s"] or 0

    by_type = (pays_qs.values("service_type")
               .annotate(total=Sum("amount_total"), n=Count("id"))
               .order_by("-total"))

    # Dépenses
    exp_qs    = Expense.objects.filter(spent_at__gte=date_from, spent_at__lte=date_to).select_related("created_by")
    exp_total = exp_qs.aggregate(s=Sum("amount"))["s"] or 0
    exp_by_cat = (exp_qs.values("category")
                  .annotate(total=Sum("amount"), n=Count("id"))
                  .order_by("-total"))

    # Achats stock
    purchases_qs = Purchase.objects.filter(purchased_at__gte=date_from, purchased_at__lte=date_to).select_related("supplier","created_by")
    pur_total = purchases_qs.aggregate(s=Sum("total_amount"))["s"] or 0

    # Salaires
    sal_qs    = Salary.objects.filter(month__gte=date_from, month__lte=date_to).select_related("employee")
    sal_total = sal_qs.aggregate(s=Sum("amount"))["s"] or 0

    # Accouchements
    acc_qs = AccouchementDetail.objects.filter(
        payment__paid_at__date__gte=date_from,
        payment__paid_at__date__lte=date_to
    ).select_related("payment__patient","payment__doctor")
    acc_total = sum(float(a.total_frais()) for a in acc_qs)
    nb_nuits_total = sum(a.nb_nuits for a in acc_qs)

    charges_total = exp_total + sal_total + pur_total
    balance = pay_total - charges_total

    return render(request, "reports/index.html", {
        "today": today,
        "date_from": date_from, "date_to": date_to,
        "pay_total": pay_total, "pay_chifa": pay_chifa, "pay_mil": pay_mil,
        "exp_total": exp_total, "pur_total": pur_total, "sal_total": sal_total,
        "charges_total": charges_total, "balance": balance,
        "by_type": by_type, "exp_by_cat": exp_by_cat,
        "pays_qs": pays_qs.order_by("-paid_at")[:50],
        "exp_qs": exp_qs.order_by("-spent_at")[:30],
        "purchases_qs": purchases_qs.order_by("-purchased_at")[:20],
        "sal_qs": sal_qs.order_by("-month")[:20],
        "acc_qs": acc_qs[:20],
        "acc_total": acc_total, "nb_nuits_total": nb_nuits_total,
    })

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def reports_export_excel(request):
    from io import BytesIO
    from openpyxl import Workbook

    today = timezone.localdate()
    month_start = today.replace(day=1)

    pay_month = Payment.objects.filter(paid_at__date__gte=month_start).aggregate(s=Sum("amount_total"))["s"] or 0
    exp_month = Expense.objects.filter(spent_at__gte=month_start).aggregate(s=Sum("amount"))["s"] or 0
    balance = pay_month - exp_month

    by_type = (Payment.objects.filter(paid_at__date__gte=month_start)
               .values("service_type")
               .annotate(total=Sum("amount_total"), n=Count("id"))
               .order_by("-total"))

    wb = Workbook()

    ws = wb.active
    ws.title = "Résumé"
    ws.append(["Période", month_start.strftime("%Y-%m-%d") + " → " + today.strftime("%Y-%m-%d")])
    ws.append(["Chiffre d'affaires", float(pay_month)])
    ws.append(["Dépenses", float(exp_month)])
    ws.append(["Bénéfice net", float(balance)])

    ws2 = wb.create_sheet("Par type")
    ws2.append(["Type", "Nombre", "Total"])
    for row in by_type:
        ws2.append([row["service_type"], int(row["n"]), float(row["total"] or 0)])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="finances_stats.xlsx"'
    return resp


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def reports_export_pdf(request):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    today = timezone.localdate()
    month_start = today.replace(day=1)

    pay_month = Payment.objects.filter(paid_at__date__gte=month_start).aggregate(s=Sum("amount_total"))["s"] or 0
    exp_month = Expense.objects.filter(spent_at__gte=month_start).aggregate(s=Sum("amount"))["s"] or 0
    balance = pay_month - exp_month

    by_type = (Payment.objects.filter(paid_at__date__gte=month_start)
               .values("service_type")
               .annotate(total=Sum("amount_total"), n=Count("id"))
               .order_by("-total"))

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    y = h - 40
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, "Finances & Statistiques")
    y -= 22

    c.setFont("Helvetica", 10)
    c.drawString(40, y, f"Période: {month_start.strftime('%Y-%m-%d')} → {today.strftime('%Y-%m-%d')}")
    y -= 18
    c.drawString(40, y, f"Chiffre d'affaires: {pay_month} DA")
    y -= 14
    c.drawString(40, y, f"Dépenses: {exp_month} DA")
    y -= 14
    c.drawString(40, y, f"Bénéfice net: {balance} DA")
    y -= 22

    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, "Détail par type")
    y -= 16

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Type")
    c.drawRightString(w - 140, y, "Nombre")
    c.drawRightString(w - 40, y, "Total")
    y -= 14
    c.setFont("Helvetica", 9)

    for r in by_type:
        if y < 60:
            c.showPage()
            y = h - 40
            c.setFont("Helvetica", 9)
        c.drawString(40, y, str(r["service_type"])[:28])
        c.drawRightString(w - 140, y, str(r["n"]))
        c.drawRightString(w - 40, y, str(r["total"]))
        y -= 12

    c.showPage()
    c.save()
    bio.seek(0)

    resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="finances_stats.pdf"'
    return resp

# Admin in-app simple settings
@login_required
@require_groups(GROUP_ADMIN,)
def settings_view(request):
    packs = LabPack.objects.prefetch_related("tests").all()
    tests = LabTest.objects.all()
    tarifs = TarifAccouchement.get()

    from .models import MedicamentConsommable
    medicaments = MedicamentConsommable.objects.all()

    if request.method == "POST" and "add_med" in request.POST:
        nom = request.POST.get("med_nom","").strip()
        prix = request.POST.get("med_prix","0").strip()
        if nom:
            try:
                MedicamentConsommable.objects.get_or_create(nom=nom, defaults={"prix_unit": float(prix) if prix else 0})
                messages.success(request, f"Médicament «{nom}» ajouté.")
            except Exception:
                messages.error(request, "Erreur lors de l'ajout.")
        return redirect("settings_view")

    if request.method == "POST" and "del_med" in request.POST:
        mid = request.POST.get("med_id","")
        try:
            MedicamentConsommable.objects.filter(pk=int(mid)).delete()
            messages.success(request, "Médicament supprimé.")
        except Exception: pass
        return redirect("settings_view")

    if request.method == "POST" and "save_tarifs" in request.POST:
        fields = ["salle","sejour_nuit","anesthesie","sage_femme","frais_admin",
                  "certificat_naissance","frais_chifa","medicaments_defaut"]
        for f in fields:
            val = request.POST.get(f,"0").strip()
            try:
                setattr(tarifs, f, float(val) if val else 0)
            except:
                pass
        tarifs.save()
        messages.success(request, "Tarifs mis à jour.")
        return redirect("settings_view")

    return render(request, "admin/settings.html", {"packs": packs, "tests": tests, "tarifs": tarifs, "medicaments": medicaments})


@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def purchase_delivery_note(request, pk:int):
    p = get_object_or_404(Purchase.objects.select_related("supplier","created_by","approved_by"), pk=pk)
    items = p.items.all()
    return render(request, "stock/purchase_delivery_note.html", {"p": p, "items": items})


def _recalc_purchase_total(purchase):
    total = 0
    for it in purchase.items.all():
        total += float(it.line_total)
    purchase.total_amount = total
    purchase.save(update_fields=["total_amount"])

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def suppliers_list(request):
    q = request.GET.get("q","").strip()
    qs = Supplier.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    return render(request, "stock/suppliers_list.html", {"suppliers": qs, "q": q})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def supplier_new(request):
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        name = request.POST.get("name","").strip()
        phone = request.POST.get("phone","").strip()
        address = request.POST.get("address","").strip()
        is_active = request.POST.get("is_active") == "1"
        if not name: errors["name"] = "Nom obligatoire"
        if not errors:
            try:
                Supplier.objects.create(name=name, phone=phone, address=address, is_active=is_active)
                messages.success(request, "Fournisseur ajouté.")
                return redirect("suppliers_list")
            except: errors["name"] = "Ce nom existe déjà."
    return render(request, "stock/supplier_form.html", {"errors": errors, "vals": vals, "title": "Nouveau fournisseur"})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def supplier_edit(request, pk:int):
    sup = get_object_or_404(Supplier, pk=pk)
    errors = {}
    if request.method == "POST":
        name = request.POST.get("name","").strip()
        phone = request.POST.get("phone","").strip()
        address = request.POST.get("address","").strip()
        is_active = request.POST.get("is_active") == "1"
        if not name: errors["name"] = "Nom obligatoire"
        if not errors:
            sup.name=name; sup.phone=phone; sup.address=address; sup.is_active=is_active
            sup.save()
            messages.success(request, "Fournisseur modifié.")
            return redirect("suppliers_list")
    return render(request, "stock/supplier_form.html", {"errors": errors, "sup": sup, "title": "Modifier fournisseur"})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def purchases_list(request):
    status = request.GET.get("status","").strip()
    qs = Purchase.objects.select_related("supplier","created_by","approved_by").all()
    if status:
        qs = qs.filter(status=status)
    return render(request, "stock/purchases_list.html", {"purchases": qs, "status": status, "statuses": PurchaseStatus})

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def purchase_new(request):
    form = PurchaseForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        p = form.save(commit=False)
        p.created_by = request.user
        p.save()
        messages.success(request, "Achat créé. Ajoutez les lignes.")
        return redirect("purchase_edit", pk=p.pk)
    suppliers = Supplier.objects.filter(is_active=True).order_by("name")
    return render(request, "stock/purchase_form.html", {
        "form": form, "title": "Nouvel achat",
        "suppliers": suppliers, "statuses": PurchaseStatus.choices,
    })

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def purchase_edit(request, pk:int):
    p = get_object_or_404(Purchase, pk=pk)
    form = PurchaseForm(request.POST or None, instance=p)
    item_form = PurchaseItemForm(request.POST or None, prefix="it")
    if request.method == "POST":
        if "save_header" in request.POST and form.is_valid():
            form.save()
            messages.success(request, "Achat mis à jour.")
            return redirect("purchase_edit", pk=pk)
        if "add_item" in request.POST:
            label = (request.POST.get("it-label") or "").strip()
            qty_str = (request.POST.get("it-qty") or "1").strip()
            price_str = (request.POST.get("it-unit_price") or "0").strip()
            if label:
                try:
                    qty = float(qty_str) if qty_str else 1
                    price = float(price_str) if price_str else 0
                    PurchaseItem.objects.create(purchase=p, label=label, qty=qty, unit_price=price)
                    _recalc_purchase_total(p)
                    messages.success(request, "Article ajouté.")
                except Exception as e:
                    messages.error(request, str(e))
            return redirect("purchase_edit", pk=pk)
        if "approve" in request.POST:
            p.status = PurchaseStatus.APPROVED
            p.approved_at = timezone.now()
            p.approved_by = request.user
            p.save(update_fields=["status","approved_at","approved_by"])
            messages.success(request, "Achat validé.")
            return redirect("purchases_list")
    items = p.items.all()
    return render(request, "stock/purchase_edit.html", {"form": form, "item_form": item_form, "purchase": p, "items": items})

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def purchase_delete_item(request, pk:int):
    it = get_object_or_404(PurchaseItem, pk=pk)
    pid = it.purchase_id
    it.delete()
    p = Purchase.objects.get(pk=pid)
    _recalc_purchase_total(p)
    return redirect("purchase_edit", pk=pid)

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def purchase_delivery_note(request, pk:int):
    p = get_object_or_404(Purchase.objects.select_related("supplier","created_by","approved_by"), pk=pk)
    items = p.items.all()
    return render(request, "stock/purchase_delivery_note.html", {"p": p, "items": items})

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def salaries_dashboard(request):
    import datetime
    today = datetime.date.today()
    # Month navigation
    month_param = request.GET.get("month") or ""
    if month_param:
        try:
            parts = month_param.split("-")
            month_start = datetime.date(int(parts[0]), int(parts[1]), 1)
        except:
            month_start = datetime.date(today.year, today.month, 1)
    else:
        month_start = datetime.date(today.year, today.month, 1)
    # prev/next month
    if month_start.month == 1:
        prev_month = datetime.date(month_start.year - 1, 12, 1)
    else:
        prev_month = datetime.date(month_start.year, month_start.month - 1, 1)
    if month_start.month == 12:
        next_month = datetime.date(month_start.year + 1, 1, 1)
    else:
        next_month = datetime.date(month_start.year, month_start.month + 1, 1)
    qs = Salary.objects.select_related("employee").filter(month=month_start)
    total = qs.aggregate(s=Sum("amount"))["s"] or 0
    employees_count = Employee.objects.filter(is_active=True).count()
    paid_count = qs.exclude(paid_at=None).count()
    unpaid_count = qs.filter(paid_at=None).count()
    return render(request, "hr/salaries_dashboard.html", {
        "salaries": qs, "total": total, "month": month_start,
        "employees": employees_count,
        "paid_count": paid_count, "unpaid_count": unpaid_count,
        "prev_month": prev_month, "next_month": next_month,
    })

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def salary_new(request):
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        employee_id = request.POST.get("employee_id","").strip()
        month = request.POST.get("month","").strip()
        amount = request.POST.get("amount","").strip()
        paid_at = request.POST.get("paid_at","").strip() or None
        note = request.POST.get("note","").strip()
        if not employee_id: errors["employee_id"] = "Choisir un employé"
        if not month: errors["month"] = "Mois obligatoire"
        if not amount: errors["amount"] = "Montant obligatoire"
        if not errors:
            try:
                # Convert YYYY-MM (from <input type=month>) to YYYY-MM-01
                if month and len(month) == 7: month = month + '-01'
                s = Salary.objects.create(employee_id=int(employee_id), month=month,
                    amount=float(amount), paid_at=paid_at or None, note=note, created_by=request.user)
                messages.success(request, "Salaire ajouté.")
                return redirect("salaries_dashboard")
            except Exception as ex:
                errors["general"] = str(ex)
    employees = Employee.objects.filter(is_active=True).order_by("full_name")
    return render(request, "hr/salary_form.html", {
        "errors": errors, "vals": vals, "employees": employees,
    })

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def employees_list(request):
    qs = Employee.objects.all()
    stats = {
        "total": qs.count(),
        "active": qs.filter(is_active=True).count(),
        "inactive": qs.filter(is_active=False).count(),
    }
    return render(request, "hr/employees_list.html", {"employees": qs, "stats": stats})

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def employee_new(request):
    errors = {}
    vals = {}
    if request.method == "POST":
        vals = request.POST
        full_name = request.POST.get("full_name","").strip()
        role = request.POST.get("role","").strip()
        phone = request.POST.get("phone","").strip()
        is_active = request.POST.get("is_active") == "1"
        if not full_name: errors["full_name"] = "Nom obligatoire"
        if not errors:
            try:
                Employee.objects.create(full_name=full_name, role=role, phone=phone, is_active=is_active)
                messages.success(request, "Employé ajouté.")
                return redirect("employees_list")
            except Exception as ex:
                errors["full_name"] = "Ce nom existe déjà."
    return render(request, "hr/employee_form.html", {"errors": errors, "vals": vals})


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def salary_edit(request, pk):
    s = get_object_or_404(Salary, pk=pk)
    errors = {}
    if request.method == "POST":
        if "delete" in request.POST:
            s.delete()
            messages.success(request, "Salaire supprimé.")
            return redirect("salaries_dashboard")
        employee_id = request.POST.get("employee_id","").strip()
        month       = request.POST.get("month","").strip()
        amount      = request.POST.get("amount","").strip()
        paid_at     = request.POST.get("paid_at","").strip() or None
        note        = request.POST.get("note","").strip()
        if not amount: errors["amount"] = "Montant obligatoire"
        if not errors:
            try:
                if month and len(month) == 7: month = month + "-01"
                s.amount  = float(amount)
                s.paid_at = paid_at or None
                s.note    = note
                if month: s.month = month
                if employee_id: s.employee_id = int(employee_id)
                s.save()
                messages.success(request, "Salaire corrigé.")
                return redirect("salaries_dashboard")
            except Exception as ex:
                errors["general"] = str(ex)
    employees = Employee.objects.filter(is_active=True).order_by("full_name")
    return render(request, "hr/salary_form.html", {
        "errors": errors, "salary": s, "employees": employees,
        "edit": True,
    })


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def employee_edit(request, pk):
    emp = get_object_or_404(Employee, pk=pk)
    errors = {}
    if request.method == "POST":
        if "delete" in request.POST:
            emp.delete()
            messages.success(request, "Employé supprimé.")
            return redirect("employees_list")
        full_name = request.POST.get("full_name","").strip()
        role      = request.POST.get("role","").strip()
        phone     = request.POST.get("phone","").strip()
        is_active = request.POST.get("is_active") == "1"
        if not full_name: errors["full_name"] = "Nom obligatoire"
        if not errors:
            try:
                emp.full_name = full_name
                emp.role      = role
                emp.phone     = phone
                emp.is_active = is_active
                emp.save()
                messages.success(request, "Employé modifié.")
                return redirect("employees_list")
            except Exception as ex:
                errors["full_name"] = "Ce nom existe déjà."
    return render(request, "hr/employee_form.html", {
        "errors": errors, "emp": emp, "edit": True,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  DOSSIERS MÉDICAUX MATERNITÉ
# ═══════════════════════════════════════════════════════════════════════════════
from .models import DossierMaternite, ExamenMaternite
from .forms  import DossierMaterniteForm, ExamenMaterniteForm

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def dossier_list(request):
    q = (request.GET.get("q") or "").strip()
    t = (request.GET.get("type") or "").strip()
    qs = DossierMaternite.objects.select_related("patient")
    if q:
        qs = (
            qs.filter(numero__icontains=q)
            | qs.filter(nom__icontains=q)
            | qs.filter(prenom__icontains=q)
            | qs.filter(patient__phone__icontains=q)
        )
    if t:
        qs = qs.filter(type_dossier=t)
    return render(request, "dossiers/list.html", {"dossiers": qs[:200], "q": q})


@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def dossier_new(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    auto_numero = _next_dossier_numero()
    form = DossierMaterniteForm(request.POST or None, initial={
        "nom":           patient.last_name,
        "prenom":        patient.first_name,
        "date_naissance":patient.birth_date,
        "numero":        auto_numero,
    })
    if request.method == "POST" and form.is_valid():
        d = form.save(commit=False)
        # If numero left blank, use auto
        if not d.numero:
            d.numero = auto_numero
        d.patient    = patient
        d.created_by = request.user
        d.save()
        AuditLog.objects.create(actor=request.user, action="CREATE", entity="DossierMaternite",
                                entity_id=str(d.id), message=str(d))
        messages.success(request, f"Dossier {d.numero} créé.")
        return redirect("dossier_detail", pk=d.id)
    return render(request, "dossiers/form.html", {
        "form": form, "patient": patient, "title": "Nouveau dossier maternité",
        "auto_numero": auto_numero,
    })


@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def dossier_detail(request, pk):
    d = get_object_or_404(DossierMaternite.objects.select_related("patient"), pk=pk)
    examens = d.patient.examens.all()[:20]
    return render(request, "dossiers/detail.html", {"d": d, "examens": examens})


@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def dossier_edit(request, pk):
    d       = get_object_or_404(DossierMaternite, pk=pk)
    form    = DossierMaterniteForm(request.POST or None, instance=d)
    if request.method == "POST" and form.is_valid():
        form.save()
        AuditLog.objects.create(actor=request.user, action="EDIT", entity="DossierMaternite",
                                entity_id=str(d.id), message=str(d))
        messages.success(request, "Dossier mis à jour.")
        return redirect("dossier_detail", pk=d.id)
    return render(request, "dossiers/form.html", {
        "form": form, "patient": d.patient,
        "title": f"Modifier dossier {d.numero}",
        "auto_numero": d.numero,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  EXAMENS MATERNITÉ
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def examen_list(request):
    q = (request.GET.get("q") or "").strip()
    t = request.GET.get("type") or ""
    qs = ExamenMaternite.objects.select_related("patient")
    if q:
        qs = (
            qs.filter(numero_manuel__icontains=q)
            | qs.filter(patient__last_name__icontains=q)
            | qs.filter(patient__first_name__icontains=q)
        )
    if t:
        qs = qs.filter(type_examen=t)
    from .models import TypeExamen
    return render(request, "dossiers/examen_list.html", {
        "examens": qs[:200], "q": q, "t": t, "TypeExamen": TypeExamen,
    })


@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def examen_new(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    # Default type
    default_type = request.GET.get("type", "ANAMNESE")
    auto_numero  = _next_examen_numero(default_type)
    form = ExamenMaterniteForm(request.POST or None, initial={
        "numero_manuel": auto_numero,
        "type_examen":   default_type,
    })
    if request.method == "POST" and form.is_valid():
        e = form.save(commit=False)
        if not e.numero_manuel:
            e.numero_manuel = _next_examen_numero(e.type_examen)
        e.patient    = patient
        e.created_by = request.user
        e.save()
        AuditLog.objects.create(actor=request.user, action="CREATE", entity="ExamenMaternite",
                                entity_id=str(e.id), message=str(e))
        messages.success(request, f"Examen {e.numero_manuel} enregistré.")
        return redirect("examen_detail", pk=e.id)
    return render(request, "dossiers/examen_form.html", {
        "form": form, "patient": patient, "title": "Nouvel examen maternité",
        "auto_numero": auto_numero,
    })


@login_required
@require_groups(GROUP_RECEPTION, GROUP_GERANT, GROUP_ADMIN)
def examen_detail(request, pk):
    e = get_object_or_404(ExamenMaternite.objects.select_related("patient"), pk=pk)
    return render(request, "dossiers/examen_detail.html", {"e": e})


@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def examen_edit(request, pk):
    e    = get_object_or_404(ExamenMaternite, pk=pk)
    form = ExamenMaterniteForm(request.POST or None, instance=e)
    if request.method == "POST" and form.is_valid():
        form.save()
        AuditLog.objects.create(actor=request.user, action="EDIT", entity="ExamenMaternite",
                                entity_id=str(e.id), message=str(e))
        messages.success(request, "Examen mis à jour.")
        return redirect("examen_detail", pk=e.id)
    return render(request, "dossiers/examen_form.html", {
        "form": form, "patient": e.patient, "title": f"Modifier examen {e.numero_manuel}"
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  SORTIE — Compléter dossier à la sortie
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def dossier_sortie(request, pk):
    """Formulaire rapide de sortie : diagnostic + date intervention + date sortie + nb nuits."""
    d = get_object_or_404(DossierMaternite.objects.select_related("patient"), pk=pk)
    errors = {}

    if request.method == "POST":
        diagnostic        = request.POST.get("diagnostic", "").strip()
        date_intervention = request.POST.get("date_intervention", "").strip() or None
        date_sortie       = request.POST.get("date_sortie", "").strip() or None
        date_depot        = request.POST.get("date_depot", "").strip() or None
        notes             = request.POST.get("notes", "").strip()

        if not date_sortie:
            errors["date_sortie"] = "Date de sortie obligatoire"

        if not errors:
            d.diagnostic        = diagnostic
            d.date_intervention = date_intervention or None
            d.date_sortie       = date_sortie
            d.date_depot        = date_depot or None
            if notes:
                d.notes = notes
            d.save()
            AuditLog.objects.create(
                actor=request.user, action="SORTIE", entity="DossierMaternite",
                entity_id=str(d.id),
                message=f"Sortie {date_sortie} — {diagnostic}"
            )
            messages.success(request, f"Dossier {d.numero} complété — sortie enregistrée.")
            # Redirect to caisse pre-filled with this patient
            return redirect(f"/encaissements/caisse/?patient_id={d.patient_id}&dossier_id={d.id}")

    return render(request, "dossiers/sortie.html", {
        "d": d,
        "errors": errors,
    })


@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN)
def examen_completer(request, pk):
    """Complétion post-naissance d'un examen (anamnèse / partogramme / nouveau-né)."""
    e = get_object_or_404(ExamenMaternite.objects.select_related("patient"), pk=pk)
    errors = {}

    if request.method == "POST":
        # Champs communs
        e.date_examen = request.POST.get("date_examen", e.date_examen) or e.date_examen

        if e.type_examen == "ANAMNESE":
            for field in ["ddr","groupage","ta","temperature","hu","cu","etat_general",
                          "tv","spl","bcf","ercf","echo","grossesse","parite",
                          "enfants_vivants","avis_reanima","conclusion","cat"]:
                val = request.POST.get(field, "").strip()
                if val:
                    setattr(e, field, val)
        elif e.type_examen == "PARTOGRAMME":
            for field in ["sage_femme","partogramme_notes","prescription"]:
                val = request.POST.get(field, "").strip()
                if val:
                    setattr(e, field, val)
        elif e.type_examen == "NOUVEAU_NE":
            for field in ["nn_sexe","nn_poids","nn_apgar","nn_notes"]:
                val = request.POST.get(field, "").strip()
                if val:
                    setattr(e, field, val)

        notes_comp = request.POST.get("notes", "").strip()
        if notes_comp:
            e.notes = notes_comp

        e.save()
        AuditLog.objects.create(
            actor=request.user, action="COMPLETER", entity="ExamenMaternite",
            entity_id=str(e.id), message=f"Complétion {e.get_type_examen_display()}"
        )
        messages.success(request, "Examen complété.")
        return redirect("examen_detail", pk=e.id)

    return render(request, "dossiers/examen_completer.html", {"e": e, "errors": errors})


# ═══════════════════════════════════════════════════════════════════════════════
#  API — Données patient + dossier pour la caisse
# ═══════════════════════════════════════════════════════════════════════════════

import json
from django.http import JsonResponse

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN, GROUP_GERANT)
def api_patient_dossier(request):
    """Retourne les données d'un patient + son dernier dossier maternité pour la caisse."""
    patient_id = request.GET.get("patient_id", "").strip()
    dossier_id = request.GET.get("dossier_id", "").strip()

    if not patient_id:
        return JsonResponse({"error": "patient_id requis"}, status=400)

    try:
        patient = Patient.objects.get(pk=int(patient_id))
    except Patient.DoesNotExist:
        return JsonResponse({"error": "Patient introuvable"}, status=404)

    # Récupérer le dossier (spécifique ou le plus récent)
    if dossier_id:
        try:
            dossier = DossierMaternite.objects.get(pk=int(dossier_id), patient=patient)
        except DossierMaternite.DoesNotExist:
            dossier = None
    else:
        dossier = DossierMaternite.objects.filter(patient=patient).order_by("-created_at").first()

    tarifs = TarifAccouchement.get()

    # Calculer le nombre de nuits depuis le dossier
    nb_nuits = 1
    if dossier and dossier.date_entree and dossier.date_sortie:
        from datetime import datetime
        try:
            delta = dossier.date_sortie - dossier.date_entree
            nb_nuits = max(1, delta.days)
        except Exception:
            nb_nuits = 1

    # Déterminer le type d'acte depuis le diagnostic ou le service
    type_acte = "NAT"
    if dossier and dossier.diagnostic:
        diag = dossier.diagnostic.lower()
        if "cesar" in diag or "césarienne" in diag or "ces" in diag:
            type_acte = "CES"

    # Calculer les frais
    frais = {
        "salle":                float(tarifs.salle),
        "nb_nuits":             nb_nuits,
        "tarif_nuit":           float(tarifs.sejour_nuit),
        "sejour":               nb_nuits * float(tarifs.sejour_nuit),
        "anesthesie":           float(tarifs.anesthesie),
        "sage_femme":           float(tarifs.sage_femme),
        "medicaments":          float(tarifs.medicaments_defaut),
        "frais_admin":          float(tarifs.frais_admin),
        "certificat_naissance": float(tarifs.certificat_naissance),
        "frais_chifa":          float(tarifs.frais_chifa),
    }
    total = (frais["salle"] + frais["sejour"] + frais["anesthesie"] +
             frais["sage_femme"] + frais["medicaments"] +
             frais["frais_admin"] + frais["certificat_naissance"])

    data = {
        "patient": {
            "id":          patient.id,
            "nom":         patient.last_name,
            "prenom":      patient.first_name,
            "phone":       patient.phone,
            "label":       f"{patient.last_name} {patient.first_name} — {patient.phone}",
        },
        "dossier": {
            "id":               dossier.id if dossier else None,
            "numero":           dossier.numero if dossier else None,
            "type_dossier":     dossier.type_dossier if dossier else "GENERAL",
            "diagnostic":       dossier.diagnostic if dossier else "",
            "date_entree":      str(dossier.date_entree) if dossier and dossier.date_entree else "",
            "date_sortie":      str(dossier.date_sortie) if dossier and dossier.date_sortie else "",
            "date_intervention":str(dossier.date_intervention) if dossier and dossier.date_intervention else "",
            "medecin":          dossier.medecin if dossier else "",
            "num_ass":          dossier.num_ass if dossier else "",
        } if dossier else None,
        "type_acte":  type_acte,
        "frais":      frais,
        "total":      total,
        "has_acc":    dossier is not None,
        "payer_type":  dossier.type_dossier if dossier and dossier.type_dossier in ("CHIFA","MILITAIRE") else "NORMAL",
        "num_ass":     dossier.num_ass if dossier else "",
    }
    return JsonResponse(data)


# ═══════════════════════════════════════════════════════════════════════════════
#  CAISSE INTELLIGENTE
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_groups(GROUP_RECEPTION, GROUP_ADMIN, GROUP_GERANT)
def caisse_view(request):
    """Caisse simplifiée : recherche patient → montant généré automatiquement."""
    tarifs = TarifAccouchement.get()
    doctors = Doctor.objects.all()
    patients_json = json.dumps([
        {"id": p.id, "label": f"{p.last_name} {p.first_name} — {p.phone}",
         "nom": p.last_name, "prenom": p.first_name, "phone": p.phone}
        for p in Patient.objects.order_by("last_name", "first_name")
    ])

    # Pré-remplissage depuis URL (redirect depuis sortie)
    prefill_patient_id = request.GET.get("patient_id", "")
    prefill_dossier_id = request.GET.get("dossier_id", "")

    if request.method == "POST":
        with transaction.atomic():
            patient_id    = request.POST.get("patient", "").strip()
            service_type  = request.POST.get("service_type", "").strip()
            payer_type    = request.POST.get("payer_type", "NORMAL").strip()
            tiers_montant = request.POST.get("tiers_montant", "0").strip()
            tiers_note    = request.POST.get("tiers_note", "").strip()
            doctor_id     = request.POST.get("doctor", "").strip()
            dossier_id_   = request.POST.get("dossier_id", "").strip()
            note          = request.POST.get("note", "").strip()

            def _f(k, default=0):
                v = request.POST.get(k, "").strip()
                try: return float(v) if v else float(default)
                except: return float(default)

            # Construire le AccouchementDetail depuis les champs
            nb_nuits   = int(request.POST.get("acc_nb_nuits", "1") or 1)
            tarif_nuit = _f("acc_tarif_nuit")
            salle      = _f("acc_salle")
            anest      = _f("acc_anesthesie")
            sf         = _f("acc_sage_femme")
            meds       = _f("acc_medicaments")
            fadmin     = _f("acc_frais_admin")
            cert       = _f("acc_certificat")
            fchifa     = _f("acc_frais_chifa")
            hon_med    = _f("acc_honoraires_medecin")
            acc_notes  = request.POST.get("acc_notes", "").strip()
            med_amb_id = request.POST.get("acc_medecin_ambulant", "").strip() or None

            sejour = nb_nuits * tarif_nuit
            total_frais = salle + sejour + anest + sf + meds + fadmin + cert + fchifa + hon_med

            tiers = float(tiers_montant) if tiers_montant else 0
            total_patient = max(0, total_frais - tiers)

            pay = Payment(
                patient_id    = int(patient_id),
                service_type  = service_type,
                payer_type    = payer_type,
                amount_patient      = total_patient,
                amount_third_party  = tiers,
                amount_total        = total_frais,
                third_party_note    = tiers_note,
                note          = note,
                created_by    = request.user,
                receipt_no    = _receipt_no(),
            )
            if doctor_id:
                pay.doctor_id = int(doctor_id)
            pay.save()

            if service_type in ("NAT", "CES"):
                acc_kwargs = dict(
                    payment=pay, type_acte=service_type,
                    salle=salle, nb_nuits=nb_nuits, tarif_nuit=tarif_nuit,
                    anesthesie=anest, sage_femme=sf, medicaments=meds,
                    frais_admin=fadmin, certificat_naissance=cert,
                    frais_chifa=fchifa, honoraires_medecin=hon_med, notes=acc_notes,
                )
                if med_amb_id:
                    try:
                        acc_kwargs['medecin_ambulant_id'] = int(med_amb_id)
                    except (ValueError, TypeError):
                        pass
                AccouchementDetail.objects.create(**acc_kwargs)

            AuditLog.objects.create(
                actor=request.user, action="CREATE", entity="Payment",
                entity_id=str(pay.id),
                message=f"Caisse — Reçu {pay.receipt_no} {pay.amount_total} DA"
            )
            messages.success(request, f"Encaissement enregistré. Reçu : {pay.receipt_no}")
            return redirect("payment_receipt", pk=pay.id)

    from .models import MedicamentConsommable
    medicaments = MedicamentConsommable.objects.filter(is_active=True).order_by("nom")
    medecins_ambulants = MedecinAmbulant.objects.filter(is_active=True)
    return render(request, "payments/caisse.html", {
        "tarifs": tarifs,
        "doctors": doctors,
        "patients_json": patients_json,
        "prefill_patient_id": prefill_patient_id,
        "prefill_dossier_id": prefill_dossier_id,
        "service_types": ServiceType.choices,
        "medicaments": medicaments,
        "medecins_ambulants": medecins_ambulants,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  MÉDECINS AMBULATOIRES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def medecin_ambulant_list(request):
    medecins = MedecinAmbulant.objects.prefetch_related('accouchements').all()
    data = []
    for m in medecins:
        accs = m.accouchements.select_related('payment__patient').order_by('-payment__paid_at')
        total = m.total_honoraires()
        payes = m.honoraires_payes()
        dus   = m.honoraires_dus()
        data.append({'medecin': m, 'accouchements': accs, 'total': total, 'payes': payes, 'dus': dus})
    return render(request, 'medecins_ambulants/list.html', {'data': data})


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def medecin_ambulant_new(request):
    form = MedecinAmbulantForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        m = form.save()
        messages.success(request, f"Médecin « {m.full_name} » ajouté.")
        return redirect('medecin_ambulant_detail', pk=m.pk)
    return render(request, 'medecins_ambulants/form.html', {'form': form, 'titre': 'Nouveau médecin ambulatoire'})


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def medecin_ambulant_edit(request, pk):
    m = get_object_or_404(MedecinAmbulant, pk=pk)
    form = MedecinAmbulantForm(request.POST or None, instance=m)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Médecin mis à jour.")
        return redirect('medecin_ambulant_detail', pk=m.pk)
    return render(request, 'medecins_ambulants/form.html', {'form': form, 'titre': f'Modifier — {m.full_name}', 'medecin': m})


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def medecin_ambulant_detail(request, pk):
    medecin = get_object_or_404(MedecinAmbulant, pk=pk)
    accs = medecin.accouchements.select_related('payment__patient').order_by('-payment__paid_at')
    payer_form = HonorairesPayerForm()
    return render(request, 'medecins_ambulants/detail.html', {
        'medecin': medecin,
        'accouchements': accs,
        'payer_form': payer_form,
        'total': medecin.total_honoraires(),
        'payes': medecin.honoraires_payes(),
        'dus':   medecin.honoraires_dus(),
    })


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def honoraires_mark_paid(request, acc_pk):
    """Marquer les honoraires d'un AccouchementDetail comme payés."""
    acc = get_object_or_404(AccouchementDetail, pk=acc_pk)
    if request.method == 'POST':
        form = HonorairesPayerForm(request.POST)
        if form.is_valid():
            acc.honoraires_payes = True
            acc.honoraires_payes_le = form.cleaned_data['honoraires_payes_le']
            acc.save(update_fields=['honoraires_payes', 'honoraires_payes_le'])
            messages.success(request, "Honoraires marqués comme payés.")
    medecin_pk = acc.medecin_ambulant_id or 0
    if medecin_pk:
        return redirect('medecin_ambulant_detail', pk=medecin_pk)
    return redirect('medecin_ambulant_list')


@login_required
@require_groups(GROUP_GERANT, GROUP_ADMIN)
def honoraires_mark_unpaid(request, acc_pk):
    """Annuler le paiement des honoraires d'un AccouchementDetail."""
    acc = get_object_or_404(AccouchementDetail, pk=acc_pk)
    if request.method == 'POST':
        acc.honoraires_payes = False
        acc.honoraires_payes_le = None
        acc.save(update_fields=['honoraires_payes', 'honoraires_payes_le'])
        messages.success(request, "Paiement des honoraires annulé.")
    medecin_pk = acc.medecin_ambulant_id or 0
    if medecin_pk:
        return redirect('medecin_ambulant_detail', pk=medecin_pk)
    return redirect('medecin_ambulant_list')
